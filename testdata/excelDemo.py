import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Shubham\\Documents\\DemoExcel.xlsx")
sheet = book.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
#
# sheet.cell(row=2, column=2).value = "Bhushan"
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet["A7"].value)

Dict = {}

for i in range(2, sheet.max_row + 1):

    for j in range(1, sheet.max_column + 1):
        # print(sheet.cell(row=i, column=j).value)
        # Dict["firstname"] = "shu" # same code is written below
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    print(Dict)
