import openpyxl


class HomePageData:
    test_HomePage_Data = [{"firstname": "shu", "lastname": "kal", "gender": "Male"},
                          {"firstname": "bhu", "lastname": "kal", "gender": "Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Shubham\\Documents\\DemoExcel.xlsx")
        sheet = book.active

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    # Dict["firstname"] = "shu" # same code is written below
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
